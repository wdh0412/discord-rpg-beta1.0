from openpyxl import load_workbook, Workbook
import math
import random
import datetime
import time
import asyncio

c_spsp = 2
c_name = 1
c_id = 2
c_skull = 9991
n_skull = '해골 병사'

hp_skull = 80
dd_skull = 10
db_skull = 0
dx_skull = 5
sp_skull = 5


wds = load_workbook("statDB.xlsx")
wss = wds.active

wbd = load_workbook("dungeonDB.xlsx")
wsd = wbd.active

def loadfiled():
    wbd = load_workbook("dungeonDB.xlsx")
    wsd = wbd.active
    
def savefiled():
    wbd.save("dungeonDB.xlsx")
    wbd.close()

def spspcc():
    spspc = wsd.cell(1, c_spsp).value
    return spspc

def skullis():
    skull = wsd.cell(2, 1).value
    if skull == n_skull:
        return True
    else:
        return False

def skullsheld():
    sheld = wsd.cell(1, 1).value
    if sheld == 10:
        wsd.cell(1, 1).value -= int(1)
        return 2222
    elif sheld < 10:
        return 3333



def skullattackpower(row_):
    skulldd = wsd.cell(2, 4).value
    udx = wsd.cell(row_, 6).value
    redx = udx + 100
    redxu = 100/redx
    reskullddre = skulldd*redxu
    reskulldd = math.trunc(reskullddre)
    return reskulldd

def checkmonster():
    checkmonsters = wsd.cell(1, 2).value
    return checkmonsters

def atspds(_row):
    atspd = wsd.cell(_row, 19).value
    return atspd

def checkskullstun():
    checkskullstn = wsd.cell(2, 8).value
    return checkskullstn

def warriorstun(_row):
    stun = wsd.cell(_row, 30).value
    return stun

def skullvshp(row_):
    uhp = wsd.cell(row_, 3).value
    return uhp

def skullattackrainer(row_):
    die = 444
    skulldd = wsd.cell(2, 4).value
    uhp = wsd.cell(row_, 3).value
    udx = wsd.cell(row_, 6).value
    usp = wsd.cell(row_, 7).value
    sheld = wsd.cell(row_, 16).value
    redx = udx + 100
    redxu = 100/redx
    reskullddre = skulldd*redxu
    reskulldd = math.trunc(reskullddre)
    a = random.randrange(1,200)
    
    if a <= usp: 
        return False
    elif usp < a: 
        if sheld == 10:
            if reskulldd >= uhp:
                return die
            elif reskulldd < uhp:
                wsd.cell(row_, 3).value -= int(reskulldd)
                savefiled()
                return 10
        elif sheld < 10:
            wsd.cell(row_, 16).value += int(1)
            savefiled()
            return 9
        
def skullattack(row_):
    die = 444
    skulldd = wsd.cell(2, 4).value
    uhp = wsd.cell(row_, 3).value
    udx = wsd.cell(row_, 6).value
    usp = wsd.cell(row_, 7).value
    sheld = wsd.cell(row_, 18).value
    redx = udx + 100
    redxu = 100/redx
    reskullddre = skulldd*redxu
    reskulldd = math.trunc(reskullddre)
    a = random.randrange(1,200)
    
    if a <= usp: 
        return False
    elif usp < a: 
        if sheld == 10:
            if reskulldd >= uhp:
                return die
            elif reskulldd < uhp:
                wsd.cell(row_, 3).value -= int(reskulldd)
                savefiled()
                return 10
        elif sheld < 10:
            wsd.cell(row_, 18).value += int(1)
            savefiled()
            return 9
            
def skullsmashattackis(row_):
    skulldd = wsd.cell(2, 4).value
    udx = wsd.cell(row_, 6).value
    redx = udx + 100
    redxu = 100/redx
    skullddsmash = skulldd*2.5
    reskullddre = skullddsmash*redxu
    reskulldd = math.trunc(reskullddre)
    return reskulldd

def skullsmashattack(row_):
    die = 444
    skulldd = wsd.cell(2, 4).value
    uhp = wsd.cell(row_, 3).value
    udx = wsd.cell(row_, 6).value
    usp = wsd.cell(row_, 7).value
    sheld = wsd.cell(row_, 18).value
    redx = udx + 100
    redxu = 100/redx
    skullddsmash = skulldd*2.5
    reskullddre = skullddsmash*redxu
    reskulldd = math.trunc(reskullddre)
    a = random.randrange(1,200)
    if a <= usp: 
        return False
    elif usp < a: 
        if sheld == 10:
            if reskulldd >= uhp:
                return die
            elif reskulldd < uhp:
                wsd.cell(row_, 3).value -= int(reskulldd)
                savefiled()
                return 10
        elif sheld < 10:
            wsd.cell(row_, 18).value += int(1)
            savefiled()
            return 9
        
def skullsmashattackrainer(row_):
    die = 444
    skulldd = wsd.cell(2, 4).value
    uhp = wsd.cell(row_, 3).value
    udx = wsd.cell(row_, 6).value
    usp = wsd.cell(row_, 7).value
    sheld = wsd.cell(row_, 16).value
    redx = udx + 100
    redxu = 100/redx
    skullddsmash = skulldd*2.5
    reskullddre = skullddsmash*redxu
    reskulldd = math.trunc(reskullddre)
    a = random.randrange(1,200)
    if a <= usp: 
        return False
    elif usp < a: 
        if sheld == 10:
            if reskulldd >= uhp:
                return die
            elif reskulldd < uhp:
                wsd.cell(row_, 3).value -= int(reskulldd)
                savefiled()
                return 10
        elif sheld < 10:
            wsd.cell(row_, 16).value += int(1)
            savefiled()
            return 9

def hillskullisover():
    skullhp = wsd.cell(2, 3).value
    hillhpskull = skullhp + 40
    rehillhpskull = hillhpskull - 80
    hilllingspring = 40 - rehillhpskull
    return hilllingspring            

def hillskull():
    skullhp = wsd.cell(2, 3).value
    hillhpskull = skullhp + 40
    rehillhpskull = hillhpskull - 80
    hilllingspring = 40 - rehillhpskull
    if hillhpskull > 80:
        wsd.cell(2, 3).value += int(hilllingspring)
        return 39
    elif hillhpskull <= 80:
        wsd.cell(2, 3).value += int(40)
        return 40
    
def hilllight1(row_):
    uhp = wsd.cell(row_, 3).value
    uhpfull = wsd.cell(row_, 43).value
    udb = wsd.cell(row_, 5).value
    hilldb = udb*2
    hillhplight = uhp + hilldb
    rehillhpskull = uhpfull - uhp
    if uhpfull < hillhplight:
        wsd.cell(row_, 3).value += int(rehillhpskull)
        return 39
    elif uhpfull <= hillhplight:
        wsd.cell(row_, 3).value += int(hilldb)
        return 40    


                
def warriorpassiveskull():
    b = random.randrange(1,100)
    if 10 >= b > 2:
        return 1010
    
    elif b > 10:
        return 3030
           
    elif b <= 2:
        return 2020        
        
            
def resetDatadg():
    loadfiled()

    print("유저 데이터를 삭제")

    wsd.delete_rows(1,wsd.max_row)
    savefiled()

    print("데이터 삭제 완료")
            
def skullhp():
    skullhp = wsd.cell(2, 3).value
    return skullhp

def warriorvsskulldd(row_):
    udd = wsd.cell(row_, 4).value
    skulldx = wsd.cell(2, 6).value
    redx = skulldx + 100
    redxskull = 100/redx
    iswarriordd = udd*redxskull
    warriordd = math.trunc(iswarriordd)
    return warriordd

def swardattackskull(row_):
    print('실행데미지')
    die = 444
    udd = wsd.cell(row_, 4).value
    skullhp = wsd.cell(2, 3).value
    skulldx = wsd.cell(2, 6).value
    skullsp = wsd.cell(2, 7).value
    skullsheld = wsd.cell(1, 1).value
    redx = skulldx + 100
    redxskull = 100/redx
    iswarriordd = udd*redxskull
    warriordd = math.trunc(iswarriordd)
    print(warriordd)
    a = random.randrange(1,200)
    if a <= skullsp:
        return False
    elif skullsp < a: 
        if skullsheld == 10:
            if warriordd >= skullhp:
                return die
            elif warriordd < skullhp:
                print(warriordd)
                wsd.cell(2, 3).value -= int(warriordd)
                savefiled()
                return 10
        elif skullsheld < 10:
            wsd.cell(1, 1).value += int(1)
            savefiled()
            return 9

def raineratcritcaldd(row_):
    udd = wsd.cell(row_, 4).value
    skulldx = wsd.cell(2, 6).value
    critcaldamge = wsd.cell(row_, 8).value
    redx = skulldx + 100
    redxskull = 100/redx
    israinerdd = udd*redxskull
    critcaldamage = math.trunc(israinerdd*critcaldamge)
    iscrdag = redxskull*critcaldamage
    realcritcalddd = math.trunc(iscrdag)
    return realcritcalddd          
        
def rainerattackskull(row_):
    print('실행데미지')
    die = 77444
    die = 444
    udd = wsd.cell(row_, 4).value
    skullhp = wsd.cell(2, 3).value
    skulldx = wsd.cell(2, 6).value
    skullsp = wsd.cell(2, 7).value
    critcalpasant = wsd.cell(row_, 9).value
    critcaldamge = wsd.cell(row_, 8).value
    skullsheld = wsd.cell(1, 1).value
    redx = skulldx + 100
    redxskull = 100/redx
    israinerdd = udd*redxskull
    rainerdd = math.trunc(israinerdd)
    critcaldamage = math.trunc(israinerdd*critcaldamge)
    iscrdag = redxskull*critcaldamage
    realcritcalddd = math.trunc(iscrdag)
    print(rainerdd)
    a = random.randrange(1,200)
    critcal = random.randrange(critcalpasant,100)
    if a <= skullsp:
        return False
    elif skullsp < a: 
        if skullsheld == 10:
            if critcalpasant >= critcal:
                if realcritcalddd >= skullhp:
                    return die
                elif realcritcalddd < skullhp:
                    print(realcritcalddd)
                    wsd.cell(2, 3).value -= int(realcritcalddd)
                    savefiled()
                    return 777
            elif critcalpasant < critcal:    
                if rainerdd >= skullhp:
                    return die
                elif rainerdd < skullhp:
                    print(rainerdd)
                    wsd.cell(2, 3).value -= int(rainerdd)
                    savefiled()
                    return 10
        elif skullsheld < 10:
            wsd.cell(1, 1).value += int(1)
            savefiled()
            return 9
        
def bupsaattackskull(row_):
    print('실행데미지')
    die = 444
    udd = wsd.cell(row_, 4).value
    skullhp = wsd.cell(2, 3).value
    skulldx = wsd.cell(2, 6).value
    skullsp = wsd.cell(2, 7).value
    skullsheld = wsd.cell(1, 1).value
    redx = skulldx + 100
    redxskull = 100/redx
    iswarriordd = udd*redxskull
    warriordd = math.trunc(iswarriordd)
    print(warriordd)
    a = random.randrange(1,200)
    if a <= skullsp:
        return False
    elif skullsp < a: 
        if skullsheld == 10:
            if warriordd >= skullhp:
                return die
            elif warriordd < skullhp:
                print(warriordd)
                wsd.cell(2, 3).value -= int(warriordd)
                savefiled()
                return 10
        elif skullsheld < 10:
            wsd.cell(1, 1).value += int(1)
            savefiled()
            return 9
            

def skill1bupsafire(row_):
    print('실행데미지')
    die = 444
    udb = wsd.cell(row_, 5).value
    skullhp = wsd.cell(2, 3).value
    skulldx = wsd.cell(2, 6).value
    skullsp = wsd.cell(2, 7).value
    skullsheld = wsd.cell(1, 1).value
    redx = skulldx + 100
    redxskull = 100/redx
    fireskil1 = udb*2
    iswarriordd = fireskil1*redxskull
    warriordd = math.trunc(iswarriordd)
    print(warriordd)
    a = random.randrange(1,200)
    if a <= skullsp:
        return False
    elif skullsp < a: 
        if skullsheld == 10:
            if warriordd >= skullhp:
                return die
            elif warriordd < skullhp:
                print(warriordd)
                wsd.cell(2, 3).value -= int(warriordd)
                savefiled()
                return 10
        elif skullsheld < 10:
            wsd.cell(1, 1).value += int(1)
            savefiled()
            return 9


def bupsafirelightwater(row_):
    wsd.cell(row_, 42).value



def checkdgNum():
    loadfiled()

    count = 0

    for row in range(2, wsd.max_row+1):
        if wsd.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count

    return count

def checkfirstrowdg():
    loadfiled()


    for row in range(2, wsd.max_row + 1):
        if wsd.cell(row,1).value is None:
            return row
            break

    _result = wsd.max_row+1

    savefiled()

    return _result

def checkdg(_name, _id):
    loadfiled()

    userNum = checkdgNum()
    for row in range(2, 3+userNum):
        if wsd.cell(row, c_name).value == _name and wsd.cell(row,c_id).value == hex(_id):

            savefiled()

            return True, row
            break
        else:
            print(" ")

    savefiled()

    return False, None

def checkdgd(_name, _id):
    loadfiled()
   
    userNum = checkdgNum()
    #print("등록된 유저수: ", userNum)
    #print("")

    #print("이름과 고유번호 탐색")
    #print("")
    for row in range(2, 3+userNum):
        if wsd.cell(row, c_name).value == _name and wsd.cell(row,c_id).value == hex(_id):
            #print("")
            #print(row, "번째 줄 name: ", wsd.cell(row,c_name).value)
            #print("입력된 name: ", _name)
            #print("이름과 일치 여부: ", wsd.cell(row, c_name).value == _name)

            #print(row,"번째 줄 id: ", wsd.cell(row,c_id).value)
            #print("입력된 id: ", hex(_id))
            #print("고유번호정보와 일치 여부: ", wsd.cell(row, c_id).value == hex(_id))
            #print("")
            savefiled()
            #print("등록된  값의 위치: ",  row, "번째 줄")
            #print("")
            return row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    savefiled()
    print("발견 실패")

    return None

def warrioratcl(_row):
    watcl = wsd.cell(_row, 8).value
    return watcl

def warrioratclrestart(_row):
    wsd.cell(_row, 19).value += int(1)
  

def warrioratclredelet(_row):
    wsd.cell(_row, 19).value -= int(1)
   


def skullman():
    loadfiled()
    wsd.cell(row=1, column=1, value = 10)
    wsd.cell(row=1, column=2, value = 1)
    wsd.cell(row=2, column=1, value = n_skull)
    wsd.cell(row=2, column=2, value = c_skull)
    wsd.cell(row=2, column=3, value = hp_skull)
    wsd.cell(row=2, column=4, value = dd_skull)
    wsd.cell(row=2, column=5, value = db_skull)
    wsd.cell(row=2, column=6, value = dx_skull)
    wsd.cell(row=2, column=7, value = sp_skull)
    wsd.cell(row=2, column=8, value = 1)
    wsd.cell(row=2, column=9, value = 80)

    savefiled()

def dgzup(_name, _id, u_hp, u_dd, u_db, u_dx, u_sp, atsp, sk1t, sk1cl, sk2t, sk3de, sk3cl1, sk3cl2, fnskt, fnskcl, sk2cl, sheild):
    loadfiled()
    _row = checkfirstrowdg()
    wsd.cell(row=_row, column=c_name, value=_name)
    wsd.cell(row=_row, column=c_id, value =hex(_id))
    wsd.cell(row=_row, column=3, value = u_hp)
    wsd.cell(row=_row, column=4, value = u_dd)
    wsd.cell(row=_row, column=5, value = u_db)
    wsd.cell(row=_row, column=6, value = u_dx)
    wsd.cell(row=_row, column=7, value = u_sp)
    wsd.cell(row=_row, column=8, value = atsp)
    wsd.cell(row=_row, column=9, value = sk1t)
    wsd.cell(row=_row, column=10, value = sk1cl)
    wsd.cell(row=_row, column=11, value = sk2t)
    wsd.cell(row=_row, column=12, value = sk3de)
    wsd.cell(row=_row, column=13, value = sk3cl1)
    wsd.cell(row=_row, column=14, value = sk3cl2)
    wsd.cell(row=_row, column=15, value = fnskt)
    wsd.cell(row=_row, column=16, value = fnskcl)
    wsd.cell(row=_row, column=17, value = sk2cl)
    wsd.cell(row=_row, column=18, value = sheild)
    wsd.cell(row=_row, column=19, value = 1)
    wsd.cell(row=_row, column=20, value = 1)
    wsd.cell(row=_row, column=21, value = 1)
    wsd.cell(row=_row, column=22, value = 1)
    wsd.cell(row=_row, column=23, value = 1)
    wsd.cell(row=_row, column=24, value = 1)
    wsd.cell(row=_row, column=25, value = 1)
    wsd.cell(row=_row, column=26, value = 1)
    wsd.cell(row=_row, column=27, value = 1)
    wsd.cell(row=_row, column=28, value = 1)
    wsd.cell(row=_row, column=29, value = 1)
    wsd.cell(row=_row, column=30, value = 1)
    wsd.cell(row=_row, column=31, value = u_hp)

    savefiled()



def skill1swardmanreis(row):
    s = wsd.cell(row,20).value
    return s    

def skill1swardmancltime(row):
    s = wsd.cell(row,17).value
    return s

def skill1swardmanclrego(row):
    wsd.cell(row,20).value += int(1)
    savefiled()

def skill1swardmanclreset(row):
    wsd.cell(row,20).value -= int(1)
    savefiled()


def dgrup(_name, _id, _hp, _dd, _db, _dx, _sp, _critcaldameger, _critcallsetr, _atspr, _sk1cl, _sk2cl, _sk3ti, _sk3cl, _sk4cl, _miss):
    loadfiled()
    _row = checkfirstrowdg()
    wsd.cell(row=_row, column=c_name, value=_name)
    wsd.cell(row=_row, column=c_id, value =hex(_id))
    wsd.cell(row=_row, column=3, value = _hp)
    wsd.cell(row=_row, column=4, value = _dd)
    wsd.cell(row=_row, column=5, value = _db)
    wsd.cell(row=_row, column=6, value = _dx)
    wsd.cell(row=_row, column=7, value = _sp)
    wsd.cell(row=_row, column=8, value = _critcaldameger)
    wsd.cell(row=_row, column=9, value = _critcallsetr)
    wsd.cell(row=_row, column=10, value = _atspr)
    wsd.cell(row=_row, column=11, value = _sk1cl)
    wsd.cell(row=_row, column=12, value = _sk2cl)
    wsd.cell(row=_row, column=13, value = _sk3ti)
    wsd.cell(row=_row, column=14, value = _sk3cl)
    wsd.cell(row=_row, column=15, value = _sk4cl)
    wsd.cell(row=_row, column=16, value = _miss)
    wsd.cell(row=_row, column=17, value = 1)
    wsd.cell(row=_row, column=18, value = 1)
    wsd.cell(row=_row, column=19, value = 1)
    wsd.cell(row=_row, column=20, value = 1)
    wsd.cell(row=_row, column=21, value = 1)
    wsd.cell(row=_row, column=22, value = 1)
    wsd.cell(row=_row, column=23, value = _hp)

    savefiled()


    print("데이터 추가 완료")

def skill1rainer(_row):
    missshift = wsd.cell(_row,16).value
    if missshift < 10:
        return 0000
    elif missshift == 10:
        wsd.cell(_row,16).value -= int(3)
        savefiled()
        return 3333

def skill1rainerreis(_row):
    s = wsd.cell(_row, 18).value
    return s


def skill1rainercl(_row):
    wsd.cell(_row, 11).value

def skill1rainerrego(_row):
    wsd.cell(_row, 18).value += int(1)
    savefiled()

def skill1rainerreset(_row):
    wsd.cell(_row, 18).value -= int(1)
    savefiled()
    

def bupsastun(_row):
    bupstun = wsd.cell(_row, 41).value
    return bupstun

def rainerstun(_row):
    inerstun = wsd.cell(_row, 22).value
    return inerstun

def raineratsp(_row):
    inerstun = wsd.cell(_row, 10).value
    return inerstun

def bupsaatsprego(_row):
    bupsarego = wsd.cell(_row, 27).value
    return bupsarego

def bupsaatsp(_row):
    bupsarego = wsd.cell(_row, 8).value
    return bupsarego

def raineratrego(_row):
    ineratsprego = wsd.cell(_row, 17).value
    return ineratsprego

def raineratgo(_row):
    wsd.cell(_row, 17).value += int(1)
    savefiled()

def bupsaatgo(_row):
    wsd.cell(_row, 27).value += int(1)
    savefiled()

def raineratreset(_row):
    wsd.cell(_row, 17).value -= int(1)
    savefiled()

def bupsaatreset(_row):
    wsd.cell(_row, 27).value -= int(1)
    savefiled()
    

def bupsaskill1clfire(row_):
    s = wsd.cell(row_, 9).value
    return s

def bupsaskill1clfirereis(row_):
    s = wsd.cell(row_, 28).value    
    return s

def bupsaskill1firerego(row_):
    wsd.cell(row_, 28).value += int(1)
    savefiled()

def bupsaskill1firereset(row_):
    wsd.cell(row_, 28).value -= int(1)
    savefiled()

def bupsaskill1cllight(row_):
    s = wsd.cell(row_, 10).value
    return s

def bupsaskill1cllightreis(row_):
    s = wsd.cell(row_, 29).value  
    return s  

def bupsaskill1llightrego(row_):
    wsd.cell(row_, 29).value += int(1)
    savefiled()

def bupsaskill1llightreset(row_):
    wsd.cell(row_, 29).value -= int(1)
    savefiled()

def bupsaskill1clwater(row_):
    s = wsd.cell(row_, 11).value
    return s

def bupsaskill1clwaterreis(row_):
    s = wsd.cell(row_, 30).value   
    return s 


def bupsaskill1lwaterrego(row_):
    wsd.cell(row_, 30).value += int(1)
    savefiled()

def bupsaskill1lwaterreset(row_):
    wsd.cell(row_, 30).value -= int(1)
    savefiled()    


def dgbup(_name, _id, _hp, _dd, _db, _dx, _sp, _atspb, _sk1fire, _sk1light, _sk1water, _sk2cl, _sk3tifire, _sk3clfire, _sk3tilight, _sk3cllight, _sk3tiwater, _sk3clwater, _sk4tmfire, _sk4clfire, _sk4tilight, _sk4cllight, _sk4tmwater, _sk4clwater, _sheldb, _bariar):
    print("dg.py - dgup")
    loadfiled()
    _row = checkfirstrowdg()
    wsd.cell(row=_row, column=c_name, value=_name)
    wsd.cell(row=_row, column=c_id, value =hex(_id))
    wsd.cell(row=_row, column=3, value = _hp)
    wsd.cell(row=_row, column=4, value = _dd)
    wsd.cell(row=_row, column=5, value = _db)
    wsd.cell(row=_row, column=6, value = _dx)
    wsd.cell(row=_row, column=7, value = _sp)
    wsd.cell(row=_row, column=8, value = _atspb)
    wsd.cell(row=_row, column=9, value = _sk1fire)
    wsd.cell(row=_row, column=10, value = _sk1light)
    wsd.cell(row=_row, column=11, value = _sk1water)
    wsd.cell(row=_row, column=12, value = _sk2cl)
    wsd.cell(row=_row, column=13, value = _sk3tifire)
    wsd.cell(row=_row, column=14, value = _sk3clfire)
    wsd.cell(row=_row, column=15, value = _sk3tilight)
    wsd.cell(row=_row, column=16, value = _sk3cllight)
    wsd.cell(row=_row, column=17, value = _sk3tiwater)
    wsd.cell(row=_row, column=18, value = _sk3clwater)
    wsd.cell(row=_row, column=19, value = _sk4tmfire)
    wsd.cell(row=_row, column=20, value = _sk4clfire)
    wsd.cell(row=_row, column=21, value = _sk4tilight)
    wsd.cell(row=_row, column=22, value = _sk4cllight)
    wsd.cell(row=_row, column=23, value = _sk4tmwater)
    wsd.cell(row=_row, column=24, value = _sk4clwater)
    wsd.cell(row=_row, column=25, value = _sheldb)
    wsd.cell(row=_row, column=26, value = _bariar)
    wsd.cell(row=_row, column=27, value = 1)
    wsd.cell(row=_row, column=28, value = 1)
    wsd.cell(row=_row, column=29, value = 1)
    wsd.cell(row=_row, column=30, value = 1)
    wsd.cell(row=_row, column=31, value = 1)
    wsd.cell(row=_row, column=32, value = 1)
    wsd.cell(row=_row, column=33, value = 1)
    wsd.cell(row=_row, column=34, value = 1)
    wsd.cell(row=_row, column=35, value = 1)
    wsd.cell(row=_row, column=36, value = 1)
    wsd.cell(row=_row, column=37, value = 1)
    wsd.cell(row=_row, column=38, value = 1)
    wsd.cell(row=_row, column=39, value = 1)
    wsd.cell(row=_row, column=40, value = 1)
    wsd.cell(row=_row, column=41, value = 1)
    wsd.cell(row=_row, column=42, value = 1)
    wsd.cell(row=_row, column=43, value = _hp)
    print("dd")

    savefiled()


    print("데이터 추가 완료")

def dungeonbariar(row_):
    bariar = wsd.cell(row_, 26).value
    return bariar

def warriorskill1stun(row_):
    s = wsd.cell(row_, 17).value
    a = s-2

    return a    

def skullsmashattackbupsa(row_):
    die = 444
    bariar = wsd.cell(row_, 26).value
    skulldd = wsd.cell(2, 4).value
    uhp = wsd.cell(row_, 3).value
    udx = wsd.cell(row_, 6).value
    usp = wsd.cell(row_, 7).value
    sheld = wsd.cell(row_, 25).value
    redx = udx + 100
    redxu = 100/redx
    skullddsmash = skulldd*2.5
    reskullddre = skullddsmash*redxu
    reskulldd = math.trunc(reskullddre)
    a = random.randrange(1,200)
    if a <= usp: 
        return False
    elif usp < a: 
        if sheld == 10:
            if bariar == 0:
                if reskulldd >= uhp:
                    return die
                elif reskulldd < uhp:
                    wsd.cell(row_, 3).value -= int(reskulldd)
                    savefiled()
                    return 10
            elif bariar > 0:
                if reskulldd > bariar:
                    breakbariar = reskulldd-bariar
                    wsd.cell(row_, 3).value -= int(breakbariar)
                    savefiled()
                    return 3
                elif reskulldd < bariar:
                    wsd.cell(row_, 26).value -= int(reskulldd)
                    savefiled()
                    return 4
        elif sheld < 10:
            wsd.cell(row_, 25).value += int(1)
            savefiled
            return 9

def swardskill1skulldamge(row_):
    print('실행데미지')
    die = 444
    udd = wsd.cell(row_, 4).value
    skullhp = wsd.cell(2, 3).value
    skulldx = wsd.cell(2, 6).value
    redx = skulldx + 100
    redxskull = 100/redx
    iswarriordd = udd*redxskull
    warriordd = math.trunc(iswarriordd)
    skill1swardman = 5/warriordd
    skill1warriar = warriordd-skill1swardman
    skill1faint = math.trunc(skill1warriar)
    return skill1faint

def swardskill1skull(row_):
    print('실행데미지')
    die = 444
    udd = wsd.cell(row_, 4).value
    skullhp = wsd.cell(2, 3).value
    skulldx = wsd.cell(2, 6).value
    skullsp = wsd.cell(2, 7).value
    skullsheld = wsd.cell(1, 1).value
    redx = skulldx + 100
    redxskull = 100/redx
    iswarriordd = udd*redxskull
    warriordd = math.trunc(iswarriordd)
    skill1swardman = 5/warriordd
    skill1warriar = warriordd-skill1swardman
    skill1faint = math.trunc(skill1warriar)
    print(warriordd)
    a = random.randrange(1,200)
    if a <= skullsp:
        return False
    elif skullsp < a: 
        if skullsheld == 10:
            if skill1faint >= skullhp:
                return die
            elif skill1faint < skullhp:
                print(skill1faint)
                wsd.cell(2, 3).value -= int(skill1faint)
                savefiled()
                return 10
        elif skullsheld < 10:
            wsd.cell(1, 1).value += int(1)
            wsd.cell(2, 8).value += int(1)
            wsd.cell(2, 3).value -= int(skill1faint)
            wsd.cell(2, 8).value -= int(1)
            savefiled()
            return 9
        

def skill1stunff():
    wsd.cell(2, 8).value += int(1)
    savefiled()  

def skill1stunss():
    wsd.cell(2, 8).value -= int(1)  
    savefiled()    


def skill1skulllight(row_):
    print('실행데미지')
    die = 444
    udb = wsd.cell(row_, 5).value
    skullhp = wsd.cell(2, 3).value
    skulldx = wsd.cell(2, 6).value
    skullsp = wsd.cell(2, 7).value
    skullsheld = wsd.cell(1, 1).value
    redx = skulldx + 100
    redxskull = 100/redx
    iswarriordd = udb*redxskull
    warriordd = math.trunc(iswarriordd)
    skill1swardman = 3/warriordd
    skill1ligt = warriordd-skill1swardman
    skill1ligt = math.trunc(skill1ligt)
    print(warriordd)
    a = random.randrange(1,200)
    if a <= skullsp:
        return False
    elif skullsp < a: 
        if skullsheld == 10:
            if skill1ligt >= skullhp:
                return die
            elif skill1ligt < skullhp:
                print(skill1ligt)
                wsd.cell(2, 8).value += int(1)
                wsd.cell(2, 3).value -= int(skill1ligt)
                asyncio.sleep(2)
                wsd.cell(2, 8).value -= int(1)
                savefiled()
                return 10
        elif skullsheld < 10:
            wsd.cell(1, 1).value += int(1)
            savefiled()
            return 9            

def bupsaskullattack(row_):
    die = 444
    bariar = wsd.cell(row_, 26).value
    skulldd = wsd.cell(2, 4).value
    uhp = wsd.cell(row_, 3).value
    udx = wsd.cell(row_, 6).value
    usp = wsd.cell(row_, 7).value
    sheld = wsd.cell(row_, 25).value
    redx = udx + 100
    redxu = 100/redx
    reskullddre = skulldd*redxu
    reskulldd = math.trunc(reskullddre)
    a = random.randrange(1,200)
    
    if a <= usp: 
        return False
    elif usp < a: 
        if sheld == 10:
            if bariar == 0:
                if reskulldd >= uhp:
                    return die
                elif reskulldd < uhp:
                    wsd.cell(row_, 3).value -= int(reskulldd)
                    savefiled()
                    return 10
            elif bariar > 0:
                if reskulldd > bariar:
                    breakbariar = reskulldd-bariar
                    wsd.cell(row_, 3).value -= int(breakbariar)
                    wsd.cell(row_, 26).value -= int(bariar)
                    savefiled()
                    return 3
                elif reskulldd < bariar:
                    wsd.cell(row_, 26).value -= int(reskulldd)
                    savefiled()
                    return 4


        elif sheld < 10:
            wsd.cell(row_, 25).value += int(1)
            savefiled
            return 9

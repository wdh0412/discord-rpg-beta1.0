from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_hp = 3
c_dd = 4
c_db = 5
c_dx = 6
c_sp = 7
c_job = 7



wds = load_workbook("statDB.xlsx")
wss = wds.active

def loadFile():
    wds = load_workbook("statDB.xlsx")
    wss = wds.active
    

def saveFile():
    wds.save("statDB.xlsx")
    wds.close()

def statllvv(row_, sstats, statpointis):
    if sstats == 1:
        wss.cell(row_,c_hp).value += int(statpointis)
        saveFile()
    elif sstats == 2:
        wss.cell(row_,c_dd).value += int(statpointis)
        saveFile()
    elif sstats == 3:
        wss.cell(row_,c_db).value += int(statpointis)
        saveFile()
    elif sstats == 4:
        wss.cell(row_,c_dx).value += int(statpointis)
        saveFile()
    elif sstats == 5:
        wss.cell(row_,c_sp).value += int(statpointis)
        saveFile()



def checkstatNum():
    print("stat.py - checkstatNum")
    loadFile()

    count = 0

    for row in range(2, wss.max_row+1):
        if wss.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count

    return count

def checkfirstrowstat():
    print("stat.py - checkfirstrowstat")
    loadFile()

    print("첫번째 빈 곳을 탐색")

    for row in range(2, wss.max_row + 1):
        if wss.cell(row,1).value is None:
            return row
            break

    _result = wss.max_row+1

    saveFile()

    return _result

def checkstat(_name, _id):
    print("state.py - checkstat")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkstatNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", wss.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", wss.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", wss.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", wss.cell(row, c_id).value == hex(_id))
        print("")

        if wss.cell(row, c_name).value == _name and wss.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

def checkstate(_name, _id):
    print("state.py - checkstat")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkstatNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", wss.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", wss.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", wss.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", wss.cell(row, c_id).value == hex(_id))
        print("")

        if wss.cell(row, c_name).value == _name and wss.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return None

def statup(_name, _id):
    print("stat.py - statz")
    loadFile()
    _row = checkfirstrowstat()

    print("첫번째 빈곳: ", _row)
    print("")

    print("데이터 추가 시작")

    wss.cell(row=_row, column=c_name, value=_name)
    print("이름 추가 | ",  wss.cell(_row,c_name).value)
    wss.cell(row=_row, column=c_id, value =hex(_id))
    print("고유번호 추가 | ", wss.cell(_row,c_id).value)

    wss.cell(row=_row, column=c_hp, value = 80)
    print("초기 HP 설정 | hp:", wss.cell(_row,c_hp).value)
    wss.cell(row=_row, column=c_dd, value = 10)
    print("초기 dd 설정 | exp:", wss.cell(_row,c_dd).value)

    wss.cell(row=_row, column=c_db, value = 10)
    print("기본 db 설정 | ", wss.cell(_row,c_db).value)
    wss.cell(row=_row, column=c_dx, value = 5)
    print("초기 dx 설정:", wss.cell(_row,c_dx).value)

    wss.cell(row=_row, column=c_sp, value = 5)
    print("초기 sp 설정:", wss.cell(_row,c_sp).value)



    

    print("")

    saveFile()


    print("데이터 추가 완료")

def statupr(_name, _id):
    print("stat.py - statz")
    loadFile()
    _row = checkfirstrowstat()

    wss.cell(row=_row, column=c_name, value=_name)
    wss.cell(row=_row, column=c_id, value =hex(_id))
    wss.cell(row=_row, column=c_hp, value = 80)
    wss.cell(row=_row, column=c_dd, value = 10)
    wss.cell(row=_row, column=c_db, value = 10)
    wss.cell(row=_row, column=c_dx, value = 5)
    wss.cell(row=_row, column=c_sp, value = 5)
    wss.cell(row=_row, column=8, value = 1.8)
    wss.cell(row=_row, column=9, value = 15)
    wss.cell(row=_row, column=10, value = 2)
    wss.cell(row=_row, column=11, value = 12)
    wss.cell(row=_row, column=12, value = 3)
    wss.cell(row=_row, column=13, value = 3)
    wss.cell(row=_row, column=14, value = 11)
    wss.cell(row=_row, column=15, value = 15)
    wss.cell(row=_row, column=16, value = 10)


    

    print("")

    saveFile()


    print("데이터 추가 완료")

def statupb(_name, _id):
    print("stat.py - statz")
    loadFile()
    _row = checkfirstrowstat()

    wss.cell(row=_row, column=c_name, value=_name)
    wss.cell(row=_row, column=c_id, value =hex(_id))
    wss.cell(row=_row, column=c_hp, value = 80)
    wss.cell(row=_row, column=c_dd, value = 10)
    wss.cell(row=_row, column=c_db, value = 10)
    wss.cell(row=_row, column=c_dx, value = 5)
    wss.cell(row=_row, column=c_sp, value = 5)
    wss.cell(row=_row, column=8, value = 4)
    wss.cell(row=_row, column=9, value = 8)
    wss.cell(row=_row, column=10, value = 8)
    wss.cell(row=_row, column=11, value = 8)
    wss.cell(row=_row, column=12, value = 5)
    wss.cell(row=_row, column=13, value = 5)
    wss.cell(row=_row, column=14, value = 20)
    wss.cell(row=_row, column=15, value = 5)
    wss.cell(row=_row, column=16, value = 15)
    wss.cell(row=_row, column=17, value = 5)
    wss.cell(row=_row, column=18, value = 15)
    wss.cell(row=_row, column=19, value = 2)
    wss.cell(row=_row, column=20, value = 30)
    wss.cell(row=_row, column=21, value = 3.5)
    wss.cell(row=_row, column=22, value = 30)
    wss.cell(row=_row, column=23, value = 4)
    wss.cell(row=_row, column=24, value = 30)
    wss.cell(row=_row, column=25, value = 10)
    wss.cell(row=_row, column=26, value = 0)



    

    print("")

    saveFile()


    print("데이터 추가 완료")

def Deletestate(_row):
    print("user.py - DeleteAccount")
    loadFile()
    print("회원탈퇴 진행")

    print("유저 데이터 삭제")
    wss.delete_rows(_row)

    saveFile()
    
    print("회원탈퇴 완료")    

def statupz(_name, _id):
    print("stat.py - statz")
    loadFile()
    _row = checkfirstrowstat()

    wss.cell(row=_row, column=c_name, value=_name)
    wss.cell(row=_row, column=c_id, value =hex(_id))
    wss.cell(row=_row, column=c_hp, value = 80)
    wss.cell(row=_row, column=c_dd, value = 10)
    wss.cell(row=_row, column=c_db, value = 10)
    wss.cell(row=_row, column=c_dx, value = 5)
    wss.cell(row=_row, column=c_sp, value = 5)
    wss.cell(row=_row, column=8, value = 3.4)
    wss.cell(row=_row, column=9, value = 7)
    wss.cell(row=_row, column=10, value = 10)
    wss.cell(row=_row, column=11, value = 1)
    wss.cell(row=_row, column=12, value = 3)
    wss.cell(row=_row, column=13, value = 1.8)
    wss.cell(row=_row, column=14, value = 10)
    wss.cell(row=_row, column=15, value = 10)
    wss.cell(row=_row, column=16, value = 30)
    wss.cell(row=_row, column=17, value = 5)
    wss.cell(row=_row, column=18, value = 10)


    

    print("")

    saveFile()


    print("데이터 추가 완료")

def isuserfullhp(_row):
    loadFile()
    userfullhp = wss.cell(_row,c_hp).value
    return userfullhp

def statInfo(_row):
    loadFile()

    _hp = wss.cell(_row,c_hp).value
    _dd = wss.cell(_row,c_dd).value
    _db = wss.cell(_row,c_db).value
    _dx = wss.cell(_row,c_dx).value
    _sp = wss.cell(_row,c_sp).value

    print("체력: ", _hp)
    print("물뎀: ", _dd)
    print("마뎀: ", _db)
    print("방어: ", _dx)
    print("스피드: ", _sp)

    saveFile()

    return _hp, _dd, _db, _dx, _sp


def warriorInfo(_row):
    loadFile()
    _hp = wss.cell(_row,c_hp).value
    _dd = wss.cell(_row,c_dd).value
    _db = wss.cell(_row,c_db).value
    _dx = wss.cell(_row,c_dx).value
    _sp = wss.cell(_row,c_sp).value
    _atsp = wss.cell(_row, 8).value
    _sk1t = wss.cell(_row, 9).value
    _sk1cl = wss.cell(_row, 10).value
    _sk2t = wss.cell(_row, 11).value
    _sk3de = wss.cell(_row, 12).value
    _sk3cl1 = wss.cell(_row, 13).value
    _sk3cl2 = wss.cell(_row, 14).value
    _fnskt = wss.cell(_row, 5).value
    _fnskcl = wss.cell(_row, 16).value
    _sk2cl = wss.cell(_row, 17).value
    _sheild = wss.cell(_row, 18).value

    saveFile()

    return _hp, _dd, _db, _dx, _sp, _atsp, _sk1t, _sk1cl, _sk2t, _sk3de, _sk3cl1, _sk3cl2, _fnskt, _fnskcl,  _sk2cl, _sheild

def rainerInfo(_row):
    loadFile()
    _hp = wss.cell(_row,c_hp).value
    _dd = wss.cell(_row,c_dd).value
    _db = wss.cell(_row,c_db).value
    _dx = wss.cell(_row,c_dx).value
    _sp = wss.cell(_row,c_sp).value
    _critcaldameger = wss.cell(_row, 8).value
    _critcallsetr = wss.cell(_row, 9).value
    _atspr = wss.cell(_row, 10).value
    _sk1cl = wss.cell(_row, 11).value
    _sk2cl = wss.cell(_row, 12).value
    _sk3ti = wss.cell(_row, 13).value
    _sk3cl = wss.cell(_row, 14).value
    _sk4cl = wss.cell(_row, 15).value
    _miss = wss.cell(_row, 16).value

    saveFile()

    return _hp, _dd, _db, _dx, _sp, _critcaldameger, _critcallsetr, _atspr, _sk1cl, _sk2cl, _sk3ti, _sk3cl, _sk4cl, _miss


def bupsaInfo(_row):
    loadFile()
    _hp = wss.cell(_row,c_hp).value
    _dd = wss.cell(_row,c_dd).value
    _db = wss.cell(_row,c_db).value
    _dx = wss.cell(_row,c_dx).value
    _sp = wss.cell(_row,c_sp).value
    _atspb = wss.cell(_row, 8).value
    _sk1fire = wss.cell(_row, 9).value
    _sk1light = wss.cell(_row, 10).value
    _sk1water = wss.cell(_row, 11).value
    _sk2cl = wss.cell(_row, 12).value 
    _sk3tifire = wss.cell(_row, 13).value
    _sk3clfire = wss.cell(_row, 14).value
    _sk3tilight = wss.cell(_row, 15).value
    _sk3cllight = wss.cell(_row, 16).value
    _sk3tiwater = wss.cell(_row, 17).value
    _sk3clwater = wss.cell(_row, 18).value
    _sk4tmfire = wss.cell(_row, 19).value
    _sk4clfire = wss.cell(_row, 20).value
    _sk4tilight = wss.cell(_row, 21).value
    _sk4cllight = wss.cell(_row, 22).value
    _sk4tmwater = wss.cell(_row, 23).value
    _sk4clwater = wss.cell(_row, 24).value
    _sheldb = wss.cell(_row, 25).value
    _bariar = wss.cell(_row, 26).value
    saveFile()

    return _hp, _dd, _db, _dx, _sp, _atspb, _sk1fire, _sk1light, _sk1water, _sk2cl, _sk3tifire, _sk3clfire, _sk3tilight, _sk3cllight, _sk3tiwater, _sk3clwater, _sk4tmfire, _sk4clfire, _sk4tilight, _sk4cllight, _sk4tmwater, _sk4clwater, _sheldb, _bariar
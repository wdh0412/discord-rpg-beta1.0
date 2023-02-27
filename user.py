from openpyxl import load_workbook, Workbook
from state import statllvv

c_name = 1
c_id = 2
c_lvl = 3
c_money = 4
c_exp = 5
c_ching = 6
c_job = 7
c_true = 8
c_statpoint = 9

default_money = 1000
default_ching = "풋내기"
default_job = "백수"

wb = load_workbook("userDB.xlsx")
ws = wb.active

def loadFile():
    wb = load_workbook("userDB.xlsx")
    ws = wb.active
    
def saveFile():
    wb.save("userDB.xlsx")
    wb.close()

def checkUserNum():
    print("user.py - checkUserNum")
    loadFile()

    count = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count
def checkFirstRow():
    print("user.py - checkFirstRow")
    loadFile()

    print("첫번째 빈 곳을 탐색")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

    _result = ws.max_row+1

    saveFile()

    return _result

def checkUuser(_name, _id):

    loadFile()

    userNum = checkUserNum()

    for row in range(2, 3+userNum):

        if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):

            saveFile()
            return row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

def checkUser(_name, _id):
    print("user.py - checkUser")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", ws.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", ws.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", ws.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", ws.cell(row, c_id).value == hex(_id))
        print("")

        if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()
            print(row)
            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

def getMoney(_name,_row):
    print("user.py - getMoney")
    loadFile()
    result = ws.cell(_row, c_job).value

    saveFile()

    return result

def jobmit(sender_row, jobdd): 
  print("user.py - remit") 
  loadFile() 
  ws.cell(sender_row, c_job).value = jobdd
  saveFile() 

def jobrow(jobrow): 
  print("user.py - remit") 
  loadFile() 
  job_row = ws.cell(jobrow, c_job).value
  saveFile() 
  return job_row

def maxexpskull(row_):
    loadFile()
    explvl = ws.cell(row_, c_lvl).value
    a = explvl*20
    return a

def userlv(row_):
    a = ws.cell(row_, c_lvl).value
    return a


def userexp(row_):
    a = ws.cell(row_, c_exp).value
    return a

def lvlupskullman(row_):
    loadFile()
    explvl = ws.cell(row_, c_lvl).value
    a = explvl*20
    
    b = ws.cell(row_, c_exp).value
    c = b + 10
    d = a - c
    e = 10 - d
    if c < a:
        ws.cell(row_, c_exp).value += int(10)
        ws.cell(row_, c_money).value += int(2000)
        saveFile()
        return 000
    elif c == a:
        ws.cell(row_, c_lvl).value += int(1)
        ws.cell(row_, c_exp).value -= int(b)
        ws.cell(row_, c_money).value += int(2000)
        ws.cell(row_, c_statpoint).value += int(3)
        saveFile()
        return 999

    elif c > a:
        ws.cell(row_, c_lvl).value += int(1)
        ws.cell(row_, c_exp).value -= int(e)
        ws.cell(row_, c_money).value += int(2000)
        ws.cell(row_, c_statpoint).value += int(3)
        saveFile()
        return 999
        

def checkstatpoint(row_):
    a = ws.cell(row_, c_statpoint).value
    return a    

def statlvlup(row_, rowstat, sstats, statpointis):
    ws.cell(row_, c_statpoint).value -= int(statpointis)
    saveFile()
    statllvv(rowstat, sstats, statpointis)






 

def Signup(_name, _id):
    print("user.py - signup")

    loadFile()

    _row = checkFirstRow()
    print("첫번째 빈곳: ", _row)
    print("")

    print("데이터 추가 시작")

    ws.cell(row=_row, column=c_name, value=_name)
    print("이름 추가 | ",  ws.cell(_row,c_name).value)
    ws.cell(row=_row, column=c_id, value =hex(_id))
    print("고유번호 추가 | ", ws.cell(_row,c_id).value)

    ws.cell(row=_row, column=c_lvl, value = 1)
    print("초기 레벨 설정 | lvl:", ws.cell(_row,c_lvl).value)
    ws.cell(row=_row, column=c_exp, value = 0)
    print("초기 경험치 설정 | exp:", ws.cell(_row,c_exp).value)

    ws.cell(row=_row, column=c_money, value = default_money)
    print("기본 자금 지급 | ", ws.cell(_row,c_money).value)
    ws.cell(row=_row, column=c_ching, value = default_ching)
    print("초기 칭호 설정:", ws.cell(_row,c_ching).value)

    ws.cell(row=_row, column=c_job, value = default_job)
    print("초기 직업 설정:", ws.cell(_row,c_job).value)
    ws.cell(row=_row, column=8, value = 1)
    ws.cell(row=_row, column=c_statpoint, value = 0)

    

    print("")

    saveFile()

    print("데이터 추가 완료")

def indgg(_row):
    indg = ws.cell(_row, c_true).value
    return indg

def indggch(_row):
  print("user.py - remit") 
  loadFile() 
  ws.cell(_row, c_true).value += int(1)
  saveFile() 

def indggchrest(_row):
  print("user.py - remit") 
  loadFile() 
  ws.cell(_row, c_true).value -= int(1)
  saveFile() 

def DeleteAccount(_row):
    print("user.py - DeleteAccount")
    loadFile()
    print("회원탈퇴 진행")

    print("유저 데이터 삭제")
    ws.delete_rows(_row)

    saveFile()
    
    print("회원탈퇴 완료")

def userlvofskll(row_):
    ss = ws.cell(row_,3).value
    return ss

def userInfo(_row):
    loadFile()

    _lvl = ws.cell(_row,c_lvl).value
    _exp = ws.cell(_row,c_exp).value
    _money = ws.cell(_row,c_money).value
    _ching = ws.cell(_row,c_ching).value
    _job = ws.cell(_row,c_job).value

    print("레벨: ", _lvl)
    print("경험치: ", _exp)
    print("보유자산: ", _money)
    print("칭호: ", _ching)
    print("직업: ", _job)

    saveFile()

    return _lvl, _exp, _money, _ching, _job


#=========================For Test==================================
def resetData():
    loadFile()

    print("유저 데이터를 삭제")

    ws.delete_rows(2,ws.max_row)
    saveFile()

    print("데이터 삭제 완료")

def addMoney(_row, _amount):
    loadFile()

    ws.cell(_row, c_money).value += _amount

    saveFile()

def addExp(_row, _amount):
    loadFile()

    ws.cell(_row, c_exp).value += _amount

    saveFile()

def adjustlvl(_row, _amount):
    loadFile()
    
    ws.cell(_row, c_lvl).value = _amount
    ws.cell(_row, c_exp).value = 0

    saveFile